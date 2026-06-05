"""
================================================================================
 Syntax Corporation © 2026 — All Rights Reserved
--------------------------------------------------------------------------------
 Project : EBS Contract Renewal PAF — BPA Renewal Automation Agent
 Module  : deliverables/build_oci_bom_xlsx.py — OCI BOM + consumption estimator
 Version : 1.0.0      Build : 2026.06.04      Date : 2026-06-04
--------------------------------------------------------------------------------
 Emits dist/OCI_BOM_estimator.xlsx — a Syntax-branded OCI Bill of Materials plus
 a per-volume consumption estimator (License Included vs BYOL). Figures come from
 the one pricing model (pricing.py).
================================================================================
"""

from __future__ import annotations

from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

import pricing

DIST = Path(__file__).resolve().parent / "dist"
DIST.mkdir(parents=True, exist_ok=True)

NAVY = "0632A0"; NAVY_DK = "041F66"; GREEN = "3CC85A"; MIST = "F4F7FC"; LINE = "DCE4F4"
WHITE = "FFFFFF"; INK = "16203A"

H = Font(name="Georgia", bold=True, color=WHITE, size=12)
HDR = Font(name="Calibri", bold=True, color=WHITE, size=11)
B = Font(name="Calibri", color=INK, size=11)
BB = Font(name="Calibri", bold=True, color=INK, size=11)
fill_navy = PatternFill("solid", fgColor=NAVY)
fill_navydk = PatternFill("solid", fgColor=NAVY_DK)
fill_mist = PatternFill("solid", fgColor=MIST)
fill_green = PatternFill("solid", fgColor=GREEN)
thin = Side(style="thin", color=LINE)
border = Border(left=thin, right=thin, top=thin, bottom=thin)
money = '"$"#,##0'


def _title(ws, text, span):
    ws.merge_cells(f"A1:{get_column_letter(span)}1")
    c = ws["A1"]; c.value = text; c.font = H; c.fill = fill_navydk
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 30


def _hdr(ws, row, headers):
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=i, value=h)
        c.font = HDR; c.fill = fill_navy; c.border = border
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[row].height = 28


def build() -> None:
    m = pricing.model()
    wb = openpyxl.Workbook()

    # ---- Sheet 1: BOM line items -----------------------------------------
    ws = wb.active; ws.title = "OCI BOM"
    _title(ws, "Syntax · EBS Contract Renewal PAF — OCI Bill of Materials", 4)
    ws["A2"] = "Deployment: License Included (BYOL ~72% lower on ADB)"; ws["A2"].font = B
    _hdr(ws, 4, ["Line item", "Role", "Metric", "Notes"])
    bom = [
        ("Oracle AI Database 26ai (ADB)", "PAF sidecar", "ECPU + storage",
         "Hosts PAF only — no EBS data, not in the data path"),
        ("Private Agent Factory", "Agent runtime", "$0 (included)",
         "No-cost add-on to AI Database 26ai"),
        ("OCI Database Tools / Managed MCP", "EBS connectivity", "metered/included*",
         "Governed PL/SQL tools + SQL Reports; *confirm at install"),
        ("OCI Generative AI", "LLM extraction", "per-token",
         "Tool-capable model (grok-4 / gpt-5 / gpt-4o)"),
        ("OCI Document Understanding", "OCR (optional)", "per-page",
         "Only for scanned renewal quotes"),
        ("Object Storage + networking", "Storage/egress", "per-GB",
         "Quote archive, traces, audit"),
    ]
    r = 5
    for item, role, metric, notes in bom:
        for i, val in enumerate((item, role, metric, notes), 1):
            c = ws.cell(row=r, column=i, value=val); c.font = B; c.border = border
            c.alignment = Alignment(vertical="center", wrap_text=True)
        if r % 2 == 0:
            for i in range(1, 5):
                ws.cell(row=r, column=i).fill = fill_mist
        r += 1
    for col, w in zip("ABCD", (34, 18, 18, 46)):
        ws.column_dimensions[col].width = w

    # ---- Sheet 2: Consumption estimator ----------------------------------
    ws2 = wb.create_sheet("Estimator")
    _title(ws2, "OCI run-cost estimator by annual renewal volume", 5)
    ws2["A2"] = "License Included vs BYOL · derived from the engagement pricing model"
    ws2["A2"].font = B
    _hdr(ws2, 4, ["Annual volume", "Lic. Incl. / mo", "Lic. Incl. / yr",
                  "BYOL / mo", "BYOL / yr"])
    r = 5
    for t in m["oci_bom"]["tiers"]:
        vals = [f"{t['volume_per_year']:,} renewals",
                t["license_included_monthly"], t["license_included_annual"],
                t["byol_monthly"], t["byol_annual"]]
        for i, val in enumerate(vals, 1):
            c = ws2.cell(row=r, column=i, value=val); c.font = B; c.border = border
            if i >= 2:
                c.number_format = money
            c.alignment = Alignment(horizontal="center" if i == 1 else "right",
                                    vertical="center")
        r += 1
    # assumptions block
    r += 1
    ws2.cell(row=r, column=1, value="Assumptions").font = BB
    for k, v in [("ADB monthly base", f"${pricing.ADB_MONTHLY_BASE:,.0f}"),
                 ("Per-renewal OCI (OCR+LLM+ECPU+storage)", f"${pricing.PER_RENEWAL_OCI:.2f}"),
                 ("BYOL factor vs License Included", f"{pricing.BYOL_FACTOR:.0%}")]:
        r += 1
        ws2.cell(row=r, column=1, value=k).font = B
        ws2.cell(row=r, column=2, value=v).font = B
    for col, w in zip("ABCDE", (22, 16, 16, 16, 16)):
        ws2.column_dimensions[col].width = w

    # ---- Sheet 3: Commercial summary -------------------------------------
    ws3 = wb.create_sheet("Commercials")
    _title(ws3, "Implementation + managed-services summary", 4)
    _hdr(ws3, 3, ["Item", "12-month", "24-month", "36-month"])
    msr = m["managed_services"]["terms"]
    rows = [
        ("Fixed-fee implementation (one-time)",
         m["implementation"]["fixed_fee"], "", ""),
        ("Managed services / month", msr["12"]["monthly"],
         msr["24"]["monthly"], msr["36"]["monthly"]),
        ("Managed services / year", msr["12"]["annual"],
         msr["24"]["annual"], msr["36"]["annual"]),
        ("Term discount", f"{msr['12']['discount_pct']}%",
         f"{msr['24']['discount_pct']}%", f"{msr['36']['discount_pct']}%"),
    ]
    r = 4
    for label, *vals in rows:
        ws3.cell(row=r, column=1, value=label).font = BB
        ws3.cell(row=r, column=1).border = border
        for i, val in enumerate(vals, 2):
            c = ws3.cell(row=r, column=i, value=val); c.font = B; c.border = border
            if isinstance(val, (int, float)):
                c.number_format = money
            c.alignment = Alignment(horizontal="right")
        r += 1
    for col, w in zip("ABCD", (38, 16, 16, 16)):
        ws3.column_dimensions[col].width = w

    out = DIST / "OCI_BOM_estimator.xlsx"
    wb.save(out)
    print(f"  OK  {out.name} ({out.stat().st_size//1024} KB)")


if __name__ == "__main__":
    build()
